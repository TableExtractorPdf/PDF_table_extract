#
# cell_control.py
# PDF_table_extract
#
# Created by Ji-yong219 on 2021-11-18
# Last modified on 2021-11-18
#
from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
import csv, json, win32com.client as win32, os


def json_text_to_list(cells):
    table = []
    for row in cells:
        row_cells = []
        for cell in row:
            row_cells.append(cell['text'])
        table.append(row_cells)

    return table

def find_merge_cell(cells):
    
    result = []
    visited = []
    rows, cols = len(cells), len(cells[0])
    for row in range(rows):
        for col in range(cols):
            colspan, rowspan = 1, 1
            curt_cell = cells[row][col]
            if len(curt_cell['text']) >  0: #(row, col) not in visited:
                # 수직 방향 검사
                if curt_cell['vspan'] == True:
                    while True:
                        try:
                            under_cell = cells[row + rowspan][col]
                            if len(under_cell['text']) == 0 and under_cell['vspan'] == True:
                                visited.append((row + rowspan, col))
                                # cells[row + rowspan][col]['text'] = 'searched'
                                rowspan += 1
                            else:
                                break
                            
                        except IndexError as e:
                            break
                # # 수평 방향 검사
                if curt_cell['hspan'] == True:
                    while True:
                        try:
                            next_cell = cells[row][col + colspan]
                            if len(next_cell['text']) == 0 and next_cell['hspan'] == True:
                                visited.append((row, col + colspan))
                                # cells[row][col + colspan]['text'] = 'searched'
                                colspan += 1
                            else:
                                break
                        except IndexError as e:
                            break
                if colspan > 1 or rowspan > 1:
                    cell_meta = {"address": get_col_alpha(col + 1) + str(row + 1), "size": [colspan, rowspan]}
                    result.append(json.dumps(cell_meta))
    return result

def get_col_alpha(col):
    address = ''
    while col > 0:
        col, remainder = divmod(col - 1, 26)  # 26 == 알파벳 개수
        address = chr(65 + remainder)
    return address

def csv_to_xlsx(dir_path, file_path):
	wb = Workbook()
	ws = wb.active
	with open(dir_path + file_path, 'r', encoding='utf8') as f:
		next(f)
		for row in csv.reader(f):
			ws.append(row)

	my_style = NamedStyle(name="style")
	my_style.border = Border(left=Side(style='dashed'), right=Side(style='dashed'), top=Side(style='dashed'), bottom=Side(style='dashed'))
	
	my_style.font = Font(name='맑은 고딕',
				size=11,
				bold=False,
				italic=False,
				underline='none',
				strike=False,
				color='FF000000')
	wb.add_named_style(my_style)
	cols = ws.max_column
	rows = ws.max_row
	header_fill = PatternFill("solid", fgColor="667b68")
	body_fill = PatternFill("solid", fgColor="f7f9f1")
	for row in ws.iter_rows(min_row=1, max_row=rows, min_col=1, max_col=cols):
		for cell in row:
			cell.style=my_style
			if(cell.row > 1):
				cell.fill = body_fill
			else:
				cell.fill = header_fill
				cell.font = Font(b=True, color="ead253")
	xlsx_path = dir_path + "xlsx/" + file_path[:-4] + ".xlsx"
	wb.save(xlsx_path)
	wb.close
	return xlsx_path


def process_table(merge_datas, xlsx_path):
	excel = win32.Dispatch('Excel.Application')
	xlsx_path = os.path.abspath(xlsx_path)
	wb = excel.Workbooks.Open(xlsx_path)
	active_sheets = wb.Sheets.Count

	for i in range(0,active_sheets):
		ws = wb.Worksheets(i + 1)
		ws.Columns.WrapText = True
	ws = wb.Worksheets(1)
	cols = ws.UsedRange.Columns.Count
	rows = ws.UsedRange.Rows.Count
	
	for i in range(1, cols + 1):
		ws.Columns(i).ColumnWidth = 250
	for i in range(1, rows + 1):
		ws.Rows(i).RowHeight = 250
	ws.Rows.AutoFit()
	ws.Rows.VerticalAlignment = 2
	ws.Rows.HorizontalAlignment = 3
	ws.Columns.AutoFit()
	for merge_data in merge_datas:
		start_col, start_row = coordinate_from_string(merge_data['address'])
		start_col = column_index_from_string(start_col)
		end_col, end_row = start_col + (merge_data['size'][0] - 1), start_row + (merge_data['size'][1] - 1) 
		end_addr = get_column_letter(end_col) + str(end_row)
		ws.Range(merge_data['address'] + ":" + end_addr).Merge()
	
	table_end_addr = get_column_letter(cols) + str(rows)

	ws.Columns.AutoFit()

	ws.Range("A1:" + table_end_addr).BorderAround(ColorIndex = 1,Weight = 4,LineStyle = 1)
	wb.Save()
	wb.Close()